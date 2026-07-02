import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# Raw distributor data organized by town
distributors_data = {
    "Nairobi CBD": [
        "Newshakah Electricals|0722 614 399",
        "Racecourse Electricals|0722 512 261",
        "Marksy Electricals|0722 379 907",
        "Twinkid Electricals|0722 270 521",
        "Powermax General|0797 733 232 / 0718 639 547",
        "Sitima Enterprises|0722 551",
        "Uganda Electricals|0716 597 556",
        "Powermax General Merchants|0743 179 796 / 0743 263 012",
        "Sitima Enterprises|0722 551l /0706 601 731",
        "Skylar Electricals|0715 890 939",
        "Lexis Electricals|0722 323 289 / 0723 668 190",
        "Goldmax Electricals|0725 457 338 / 0720 948 715",
        "Powertex Electricals|0703 856 592 / 0726 759 780",
        "Delite Electricals|0733 615 714 / 0701 153 222",
        "Parma Electricals|0722 766 211",
        "Brima Electricals|0722 761 707",
        "Prudential East Africa|0723 802 147",
    ],
    "Nairobi Gikomba": [
        "Nafuu Classic Hardware|0709 750 411",
        "Mt. Kenya General Hardware|0709 750 411",
    ],
    "Nairobi - Peri Urban": [
        "Nationwide Electricals|0733 822 401",
    ],
    "Nairobi - Westlands": [
        "Orascom|0724 786 786",
        "Power Innovations DIY|0729 239 344",
    ],
    "Nairobi - Industrial Area": [
        "Filmico Agencies|0722 813 510",
        "Protec Electricals|0722 726 700",
        "NexPower Ltd|0715 813 911",
        "Afriwell Ltd|0724 439 646",
    ],
    "Nairobi - Karen": [
        "Sysco Systems|0722 521 284",
        "Phoenix Timber|0714 261 219",
    ],
    "Nairobi - Eastleigh": [
        "Ali Glazier|0722 522 452",
        "Hi Net Power|0726 392 045",
    ],
    "Nairobi - Utawala": [
        "Vertex Electricals|0724 681 773",
        "Jomoda Electricals|0722 491 073",
        "Alinon Electrical|0726 068 153",
        "Amani Electricals|0721 603 579",
        "Hada Electricals|0791 266 958",
    ],
    "Nairobi - Ruai": [
        "Meruvin Electrical|0721 343 990",
        "Millitech Electrical|0722 832 954",
        "Runga Holdings Ltd|0725 534 834",
    ],
    "Nairobi - Ruiru": [
        "Isso Electricals|0723 922 326",
        "Powermax Spur Mall|0739 990 014",
        "Monross Hardware|0720 250 387",
        "Three Rings Hardware|0720 839 474 / 0726 806 080",
    ],
    "Nairobi - Kitengela": [
        "Outsource R.K Limited|0722 835 799",
        "Delta Electrical|0723 391 830",
        "Kitengela Electrical|0723 766 746",
        "Gilan Electrical|0722 866 979",
        "Vemax Electricals|0718 566 393",
    ],
    "Nairobi - Kiserian": [
        "Electrical Techniques Limited|0719 626 200",
        "PMK Hardware|0715 645 669",
    ],
    "Nairobi - Kikuyu": [
        "Neagles Electrical|0723 446 245",
        "Bernley Hardware|0724 377 710",
    ],
    "Nairobi - Langata": [
        "Chinchin Electrical|0726 509 036",
        "Haraka Electrical|0722 994 172",
        "Sidis Electrical|0722 270 789",
        "Carden Electrical|0721 481 383",
    ],
    "Nairobi - Mlolongo": [
        "Lak Enterprise|0720 513 228",
        "Nemasha Electrical|0733 639 480",
        "Kaymorr Electrical|0726 383 801",
    ],
    "Nairobi - Embakasi": [
        "Steamplant|0722 309 161",
        "Nissi Hardware|0722 263 989",
    ],
    "Nairobi - Kiambu": [
        "Nafuu Classic Hardware - Thindigua|0709 750 411",
        "Nafuu Classic Hardware - Prestige|0709 750 412",
    ],
    "Nandi Hills": [
        "Modern Electricals|0726 018 100",
    ],
    "Nanyuki": [
        "G.G. Githui Hardware|0722 457 496",
        "Trunk Trail Hardware|0706 914 981",
        "Three Core PVC Hardware|0722 385 643",
        "Screwfix Hardware Ltd|0708 295 393",
    ],
    "Bomet": [
        "Simtok Hardware|0723 186 156",
        "Len's Electricals|0726 632 361",
    ],
    "Naivasha": [
        "Thamic Electricals|0722 566 849",
        "Njokama Electricals|0722 579 717",
        "Naivasha General Master|0700 094 083",
        "General Masters|0700 094 083",
        "Umeme Electrical Services|0725 444 017",
    ],
    "Nakuru": [
        "Mache Hardware|0752 269 559",
        "Waka Electrical Services|0722 675 848",
        "Jack Electricals|0722 965 907",
        "Clifflink Kenya Ltd|0711 186 869",
        "Muturi Electricals|0723 940 047 / 0715 449 826",
        "Rift Metro Hardware Store Limited|0723 084 194",
        "Pioneer Hardware|0117 514 003",
    ],
    "Eldoret": [
        "Eagles Hardware Dealers|0736 652 666",
        "Eagle Hardware Distributors Ltd|0724 629 345",
        "B.M Electrical|0721 517 685",
        "Jossin Tiles & Hardware|0722 620 149",
        "West Link Limited|0725 979 381",
        "Csoft Electrical|0720 228 582",
        "Kezzie Electrical|0721 992 707",
        "Zakir's Limited|0711 252 366",
        "Chemas Electrical|0723 297 178",
        "Angaza Electricals|0722 313 559",
    ],
    "Embu": [
        "Teisha Hardware|0722 144 544",
        "Alucob Electricals|0722 358 666",
        "Sai Electrical & Hardware|0722 942 191",
    ],
    "Gilgil": [
        "United Electricals|0720 620 026",
        "Jakenian Electricals|0722 774 812",
        "Kinkam Electrical & Plumbing|0701 893 500",
        "Technologent Electricals|0722 651 565",
        "Sameer Electricals|0711 255 947",
    ],
    "Homa Bay": [
        "Multimax Solutions Ltd|0720 948 644",
        "Levix Electricals|0724 155 236",
    ],
    "Isiolo": [
        "Sonshine Electricals|0724 284 631",
        "Antie Flow Electricals|0725 828 939",
    ],
    "Kagumo": [
        "Decira Electricals|0718 182 221",
    ],
    "Kabarnet": [
        "Stima Zone Electricals|0716 601 414",
        "Blessing Hardware|0725 286 929",
    ],
    "Kagio": [
        "Daystar Electricals|0722 915 027",
        "Customer Hardware|0722 752 638",
        "Zetro Electricals|0722 915 027",
        "Twin Kid Electricals|0714 943 955",
    ],
    "Kajiado": [
        "Sunset Electronics|0724 872 891",
    ],
    "Kakamega": [
        "Winmax Electricals|0704 392 045",
        "Changa Electricals|0722 473 480",
        "Changa Fancy|0722 485 132",
        "Mazao Multibiz|0720 551 815",
        "Kakuma Hardware|0721 796 664",
    ],
    "Kapenguria": [
        "Baviez Enterprises and General|0728 588 462",
        "Climax Electricals|0714 001 596",
        "Digital Base Electricals|0723 757 770",
        "Paraywa Electricals|0700 035 222",
    ],
    "Kapsabet": [
        "British Electrical|0712 163 085",
        "Sochi Gaa|0723 844 152",
        "Berur Electricals|0721 341 548 / 0721 423 306",
        "Pennyiel Electricals|0700 035 222",
    ],
    "Karatina": [
        "Myco Electricals|0733 962 081",
        "Fortune Electricals|0724 629 041 / 0738 254 668",
    ],
    "Kendu Bay": [
        "Jasho Hardware|0725 120 880",
    ],
    "Kenol": [
        "Peka Electricals|0712 053 602",
        "Power Beam Electricals|0726 895 009",
    ],
    "Kericho": [
        "New Dynamic Electricals|0720 700 194",
        "Embassy Crystal|0722 794 351",
        "Pulsemax Enterprises|0724 548 140",
    ],
    "Kerugoya": [
        "Digital Electricals|0722 625 529",
        "Davelite Electricals|0702 400 677",
        "Empire Electrical|0702 069 798",
    ],
    "Kitale": [
        "Suam Hardware|0735 348884",
        "Blacky Electrical|0720 442 440",
        "Dan Town Electrical|0723 683 901",
        "Shakim Electricals|0721 242 799",
        "Deniz Jumbo Electricals|0721 278 378",
        "Blessed Electricals|0111 536 210",
        "Shammah Electricals|0722 952 684",
        "Amazon Electricals|0710 441 521",
        "Sun City Electricals|0714 123 117",
        "Tranz Nzoia Hardware|0724 992 988",
    ],
    "Kisii": [
        "Michael Star General|0716 792 538",
        "Getembe Electricals|0720 905 919",
        "Getembe Prime Distributors|0723 614 056 / 0743 286 754",
        "Gucha Electricals|0729 859 295",
        "Stenna Electricals|0722 148 971",
        "Backlight Electrical|0718 367 822",
        "Highstar Electrical|0726 579 354",
        "Alhemis Ltd|0721 440 606",
    ],
    "Kisumu": [
        "Akira Electrical|0704 785 011",
        "Kens Electrical|0721 891 221",
    ],
    "Kilifi": [
        "Alego Electrical|0721 575 139",
        "Gorofani Hardware|0726 205 001",
        "Selfdream Electricals|0721 595 864",
    ],
    "Kitui": [
        "Raicha's Electro Services|0721 240 720",
        "Modern Electricals|0722 885 050",
        "Kismart Electricals|0720 129 010",
        "CityMart Electricals|0712 679 002",
    ],
    "Kithimani": [
        "Al-Madina Hardware|0722 625 529",
        "Mindfulness Supplies|0701 329 224",
        "Kilifi Electricals|0721 535 948",
    ],
    "Kipsigak Junction": [
        "Kamili Electrical|0721 380 192",
        "Best Outlet|0720 873 348",
        "Standard Citylights Ltd|0720 663 237",
        "Jutik Electricals|0701 720 673",
    ],
    "Lodwar": [
        "Tilak Hardware|0713 867 264",
        "Young Start Hardware|0700 759 090",
        "Rikinei Hardware|0723 267 267",
    ],
    "Londiani": [
        "Royal Electricals|0722 874 844",
        "Powermax Londiani|0799 253 837",
    ],
    "Machakos": [
        "New Vision Electricals|0723 361 500",
        "Fidelity Electricals|0722 277 101",
        "Hivida Electricals|0723 442 103",
    ],
    "Majengo": [
        "Tamuz Electrical|0706 993 782",
    ],
    "Malindi": [
        "Burhani Electricals|0700 281 571",
        "Mulla Electricals|0720 820 095",
        "Jafro Electricals|0724 107 661",
        "Malsat Traders|0708 878 522",
        "Mapete Hardware|0746 000 000",
    ],
    "Mau Narok": [
        "Krunal Enterprises Ltd|0701 461 122",
        "Bright Electricals|0725 913 607",
        "Mau Narok Hardware|0720 756 362",
    ],
    "Masii": [
        "Monte Sarre Electricals|0729 691 617",
        "Peka Electricals|0722 491 385",
        "Salem Electronics|0726 481 623",
        "Mwaliko Electricals|0723 880 076",
    ],
    "Maua": [
        "Mambo Electricals|0717 954 505",
    ],
    "Mbita": [
        "Hajo Electricals|0725 424 825",
        "Joy Electricals|0700 787 254",
    ],
    "Meru": [
        "Al Nuur Electricals|0724 869 653",
        "Cate Electricals|0725 314 956",
        "Favour Sheryl Electricals|0703 132 902",
        "Mwangaza Electricals|0720 840 722",
        "Chandy's Solution Ltd|0722 643 740",
        "Chirima Electricals|0721 408 532",
    ],
    "Migori": [
        "Marowa Stores|0700 088 222",
    ],
    "Mombasa": [
        "Topline Electricals|0722 768 085",
        "Quarter Capital|0722 775 338",
        "Victory Electricals|0762 951 817",
        "Tonlect Electricals|0722 230 405",
    ],
    "Moi's Bridge": [
        "Elephant Hardware and Electricals|0722 296 737",
    ],
    "Mukurweini": [
        "Asmat Enterprises|0722 816 784",
    ],
    "Mumias": [
        "Pijey Electricals Ltd|0724 599 616",
    ],
    "Muranga": [
        "Beaver Electricals|0722 283 574",
        "Wamitwe Electricals|0722 969 162",
    ],
    "Mtwapa": [
        "Chariam Electricals|0798 529 060",
        "Captain Electricals|0724 583 922",
        "Jibril Hardware|0724 236 691",
    ],
    "Mwea": [
        "Jim Electricals|0722 618 963",
        "Chantech Electricals|0721 289 082",
        "Smartech Electricals|0728 122 276",
    ],
    "Molo": [
        "J.K.M Electricals|0727 446 194",
        "Jose Electricals|0726 297 824",
    ],
    "Mwingi": [
        "Mwenge Hardware|0723 330 402",
        "Mutemi Hardware|0796 882 020",
    ],
    "Tala": [
        "Safi Electricals|0722 652 493",
        "Vision Link Electricals|0790 922 383 / 0722 599 373",
    ],
    "Thika": [
        "Jope Electricals|0722 262 493",
        "Wenton Enterprises|0721 706 765",
        "Chania Chain Supplies|0726 769 383",
    ],
    "Ugunja": [
        "Magma Electrical|0704 440 174",
    ],
    "Ukunda": [
        "Hotsun Enterprises - Beach Rd|0720 557 821",
        "Mao Hardware Mvindeni|0715 711 068",
        "Diani Hardware|0722 404 606",
    ],
    "Watamu": [
        "Mwangaza Hardware|0714 047 166",
        "Canopus Hardware|0717 108 041",
    ],
    "Webuye": [
        "Jasler Electrical|0722 643 055",
        "Undugu Electrical|0721 468 833",
    ],
    "Wote": [
        "Uniq-Point Electricals|0710 761 833",
    ],
}

# Parse data into rows
rows = []
for town, distributors in distributors_data.items():
    for distributor_entry in distributors:
        parts = distributor_entry.split('|')
        if len(parts) == 2:
            name = parts[0].strip()
            phone = parts[1].strip()
            
            rows.append({
                'Company Name': name,
                'Trade Name': '',
                'Country': 'Kenya',
                'Email': '',
                'Phone': phone,
                'Website': '',
                'City': town,
                'Address': '',
                'Latitude': '',
                'Longitude': '',
                'Type': 'distributor',
                'Status': 'active',
                'Featured': 'no',
                'Contract Start (YYYY-MM-DD)': '',
                'Contract End (YYYY-MM-DD)': '',
                'Product Categories (comma-separated)': 'Electronics',
                'Registration Number': '',
                'Notes': '',
            })

# Create DataFrame
df = pd.DataFrame(rows)

# Create Excel workbook
output_file = 'distributor_import_template.xlsx'
df.to_excel(output_file, sheet_name='Data', index=False)

# Load workbook for styling
wb = openpyxl.load_workbook(output_file)
ws_data = wb['Data']

# Define colors and styles
header_fill = PatternFill(start_color='0D6E63', end_color='0D6E63', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Style header row
for col_num, column_title in enumerate(df.columns, 1):
    cell = ws_data.cell(row=1, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border

# Set column widths and apply borders to all cells
column_widths = {
    'A': 20, 'B': 15, 'C': 12, 'D': 18, 'E': 15, 'F': 20, 'G': 12, 'H': 20,
    'I': 12, 'J': 12, 'K': 12, 'L': 12, 'M': 10, 'N': 18, 'O': 18, 'P': 25, 'Q': 15, 'R': 20
}

for col_letter, width in column_widths.items():
    ws_data.column_dimensions[col_letter].width = width

# Apply borders to all data cells
for row in ws_data.iter_rows(min_row=1, max_row=ws_data.max_row, min_col=1, max_col=len(df.columns)):
    for cell in row:
        cell.border = border
        if cell.row > 1:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Freeze header row
ws_data.freeze_panes = 'A2'

# Create Instructions sheet
ws_instructions = wb.create_sheet('Instructions')
ws_instructions.column_dimensions['A'].width = 80

instructions = [
    ['BULK DISTRIBUTOR IMPORT - TEMPLATE INSTRUCTIONS'],
    [''],
    ['REQUIRED FIELDS:'],
    ['• Company Name - Name of the distributor company'],
    ['• Country - Country name (e.g., Kenya, Uganda, Tanzania, Rwanda, Burundi, DRC, South Sudan)'],
    [''],
    ['OPTIONAL FIELDS:'],
    ['• Trade Name - Alternate business name'],
    ['• Email - Contact email address'],
    ['• Phone - Contact phone number'],
    ['• Website - Company website URL'],
    ['• City - City location'],
    ['• Address - Physical address'],
    ['• Latitude - Decimal latitude coordinate'],
    ['• Longitude - Decimal longitude coordinate'],
    ['• Type - One of: distributor, dealer, stockist, agent (default: distributor)'],
    ['• Status - One of: active, inactive, suspended (default: active)'],
    ['• Featured - yes/no or true/false (default: no)'],
    ['• Contract Start - Date in YYYY-MM-DD format'],
    ['• Contract End - Date in YYYY-MM-DD format'],
    ['• Product Categories - Comma-separated category names'],
    ['• Registration Number - Official registration number'],
    ['• Notes - Additional notes or remarks'],
    [''],
    ['NOTES:'],
    ['1. Ensure all country names match the system database exactly'],
    ['2. Date format must be YYYY-MM-DD (e.g., 2024-01-15)'],
    ['3. Latitude/Longitude should be decimal numbers (-90 to 90 for lat, -180 to 180 for long)'],
    ['4. Product categories must exist in the system'],
    ['5. Duplicate company names will be created as separate records'],
    ['6. Invalid rows will be skipped and reported after import'],
    ['7. Use commas to separate multiple product categories'],
    ['8. Email addresses must be valid email format'],
    ['9. This template contains pre-populated Kenya distributors data'],
    ['10. Modify or add entries as needed before importing'],
]

for row_idx, instruction_row in enumerate(instructions, 1):
    cell = ws_instructions.cell(row=row_idx, column=1)
    cell.value = instruction_row[0]
    if row_idx == 1:
        cell.font = Font(bold=True, size=12, color='0D6E63')

# Save workbook
wb.save(output_file)

print(f"✓ Excel file '{output_file}' created successfully!")
print(f"✓ Total distributors imported: {len(df)}")
print(f"✓ Total towns/cities: {len(distributors_data)}")
print(f"\nFile contains:")
print(f"  - Data sheet: {len(df)} distributor records")
print(f"  - Instructions sheet: Import guidelines and field descriptions")
